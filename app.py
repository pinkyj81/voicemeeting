from flask import Flask, jsonify, render_template, request, send_file
from deep_translator import GoogleTranslator
import os
import re
from io import BytesIO
from datetime import datetime

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)


def get_db_connection():
    conn_str = (
        f"DRIVER={{{os.getenv('DB_DRIVER', 'ODBC Driver 18 for SQL Server')}}};"
        f"SERVER={os.getenv('DB_SERVER', 'ms1901.gabiadb.com')};"
        f"DATABASE={os.getenv('DB_DATABASE', 'yujincast')};"
        f"UID={os.getenv('DB_USERNAME', 'pinkyj81')};"
        f"PWD={os.getenv('DB_PASSWORD', 'zoskek38!!')};"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str)


def normalize_user_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "참가자"
    return text[:40]


@app.route("/")
def lobby():
    return render_template("rooms.html")


@app.route("/meeting/<room_code>")
def meeting_page(room_code):
    if not re.fullmatch(r"\d{6}", room_code or ""):
        return "invalid room code", 400

    room_name = ""
    language = "ko-ja"
    user_name = normalize_user_name(request.args.get("name", ""))
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT TOP 1 RoomName, ISNULL(Language, 'ko-ja') AS Language
                FROM dbo.MeetingRooms
                WHERE RoomCode = ? AND IsActive = 1
                ORDER BY RoomID DESC
                """,
                room_code,
            )
            row = cursor.fetchone()
            if not row:
                return "room not found", 404
            room_name = row[0] or ""
            language = row[1] or "ko-ja"
    except Exception as exc:
        return f"database error: {exc}", 500

    # Redirect to language-specific page
    if language == "ko-en":
        return render_template(
            "index_en.html",
            room_code=room_code,
            room_name=room_name,
            user_name=user_name,
        )
    else:
        return render_template(
            "index.html",
            room_code=room_code,
            room_name=room_name,
            user_name=user_name,
        )


@app.get("/api/rooms")
def list_rooms():
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT TOP 100
                    RoomCode,
                    RoomName,
                    ISNULL(CreatedBy, '') AS CreatedBy,
                    CONVERT(VARCHAR(19), SWITCHOFFSET(CreatedAt, '+09:00'), 120) AS CreatedAt,
                    ISNULL(Language, 'ko-ja') AS Language
                FROM dbo.MeetingRooms
                WHERE IsActive = 1
                ORDER BY RoomID DESC
                """
            )
            rows = cursor.fetchall()

        rooms = [
            {
                "room_code": row[0],
                "room_name": row[1],
                "created_by": row[2],
                "created_at": row[3],
                "language": row[4],
            }
            for row in rows
        ]
        return jsonify({"rooms": rooms})
    except Exception as exc:
        return jsonify({"error": f"failed to load rooms: {exc}"}), 500


@app.get("/api/speakers")
def list_speakers():
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT TOP 200 Name
                FROM (
                    SELECT LTRIM(RTRIM(ISNULL(SpeakerName, ''))) AS Name
                    FROM dbo.MeetingMessages
                    UNION
                    SELECT LTRIM(RTRIM(ISNULL(CreatedBy, ''))) AS Name
                    FROM dbo.MeetingRooms
                ) Names
                WHERE Name <> ''
                ORDER BY Name ASC
                """
            )
            rows = cursor.fetchall()

        names = [row[0] for row in rows]
        return jsonify({"names": names})
    except Exception as exc:
        return jsonify({"error": f"failed to load names: {exc}"}), 500


@app.post("/api/rooms")
def create_room():
    payload = request.get_json(silent=True) or {}
    room_name = str(payload.get("room_name", "")).strip()
    created_by = str(payload.get("created_by", "")).strip()
    language = str(payload.get("language", "ko-ja")).strip().lower()

    if not room_name:
        return jsonify({"error": "room_name is required"}), 400

    if len(room_name) > 200:
        return jsonify({"error": "room_name is too long (max 200)"}), 400

    if len(created_by) > 100:
        return jsonify({"error": "created_by is too long (max 100)"}), 400

    if language not in {"ko-ja", "ko-en"}:
        return jsonify({"error": "language must be ko-ja or ko-en"}), 400

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO dbo.MeetingRooms (RoomName, CreatedBy, Language)
                OUTPUT inserted.RoomCode, inserted.RoomName, inserted.Language
                VALUES (?, ?, ?)
                """,
                room_name,
                (created_by or None),
                language,
            )
            row = cursor.fetchone()
            conn.commit()

        return jsonify(
            {
                "room_code": row[0],
                "room_name": row[1],
                "language": row[2],
                "entry_url": f"/meeting/{row[0]}",
            }
        )
    except Exception as exc:
        return jsonify({"error": f"failed to create room: {exc}"}), 500


@app.delete("/api/rooms")
def delete_rooms():
    payload = request.get_json(silent=True) or {}
    room_codes = payload.get("room_codes", [])

    if not isinstance(room_codes, list):
        return jsonify({"error": "room_codes must be a list"}), 400

    if not room_codes:
        return jsonify({"error": "room_codes cannot be empty"}), 400

    if len(room_codes) > 100:
        return jsonify({"error": "too many room codes (max 100)"}), 400

    # Validate all codes before attempting delete
    for code in room_codes:
        if not re.fullmatch(r"\d{6}", str(code or "")):
            return jsonify({"error": f"invalid room code: {code}"}), 400

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            # Use soft delete by setting IsActive = 0
            placeholders = ",".join(["?" for _ in room_codes])
            cursor.execute(
                f"""
                UPDATE dbo.MeetingRooms
                SET IsActive = 0
                WHERE RoomCode IN ({placeholders})
                """,
                *room_codes,
            )
            conn.commit()
            deleted_count = cursor.rowcount

        return jsonify(
            {
                "success": True,
                "deleted_count": deleted_count,
                "message": f"Deleted {deleted_count} room(s)",
            }
        )
    except Exception as exc:
        return jsonify({"error": f"failed to delete rooms: {exc}"}), 500


@app.post("/api/messages")
def save_message():
    payload = request.get_json(silent=True) or {}
    room_code = str(payload.get("room_code", "")).strip()
    speaker_name = str(payload.get("speaker_name", "")).strip()
    language_code = str(payload.get("language_code", "")).strip().lower()
    original_text = str(payload.get("original_text", "")).strip()
    translated_text_raw = payload.get("translated_text")
    translated_text = "" if translated_text_raw is None else str(translated_text_raw).strip()

    if not re.fullmatch(r"\d{6}", room_code):
        return jsonify({"error": "room_code must be 6 digits"}), 400
    if language_code not in {"ko", "ja", "en"}:
        return jsonify({"error": "language_code must be ko, ja, or en"}), 400
    if not original_text:
        return jsonify({"error": "original_text is required"}), 400
    if len(speaker_name) > 100:
        return jsonify({"error": "speaker_name is too long (max 100)"}), 400

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT TOP 1 RoomID
                FROM dbo.MeetingRooms
                WHERE RoomCode = ? AND IsActive = 1
                ORDER BY RoomID DESC
                """,
                room_code,
            )
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "room not found"}), 404

            room_id = row[0]
            cursor.execute(
                """
                INSERT INTO dbo.MeetingMessages (
                    RoomID,
                    RoomCode,
                    SpeakerName,
                    LanguageCode,
                    OriginalText,
                    TranslatedText
                )
                OUTPUT
                    inserted.MessageID,
                    CONVERT(VARCHAR(19), SWITCHOFFSET(inserted.MessageAt, '+09:00'), 120) AS MessageAt
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                room_id,
                room_code,
                (speaker_name or None),
                language_code,
                original_text,
                (translated_text or None),
            )
            inserted = cursor.fetchone()
            conn.commit()

        return jsonify(
            {
                "saved": True,
                "message_id": inserted[0] if inserted else None,
                "message_at": inserted[1] if inserted else None,
            }
        )
    except Exception as exc:
        return jsonify({"error": f"failed to save message: {exc}"}), 500


@app.get("/api/messages")
def list_messages():
    room_code = str(request.args.get("room_code", "")).strip()
    since_id_raw = str(request.args.get("since_id", "")).strip()
    if not re.fullmatch(r"\d{6}", room_code):
        return jsonify({"error": "room_code must be 6 digits"}), 400

    since_id = None
    if since_id_raw:
        if not since_id_raw.isdigit():
            return jsonify({"error": "since_id must be a positive integer"}), 400
        since_id = int(since_id_raw)

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            if since_id is None:
                cursor.execute(
                    """
                    SELECT TOP 300
                        MessageID,
                        ISNULL(SpeakerName, '') AS SpeakerName,
                        LanguageCode,
                        OriginalText,
                        ISNULL(TranslatedText, '') AS TranslatedText,
                        CONVERT(VARCHAR(19), SWITCHOFFSET(MessageAt, '+09:00'), 120) AS MessageAt
                    FROM dbo.MeetingMessages
                    WHERE RoomCode = ?
                    ORDER BY MessageID ASC
                    """,
                    room_code,
                )
            else:
                cursor.execute(
                    """
                    SELECT TOP 300
                        MessageID,
                        ISNULL(SpeakerName, '') AS SpeakerName,
                        LanguageCode,
                        OriginalText,
                        ISNULL(TranslatedText, '') AS TranslatedText,
                        CONVERT(VARCHAR(19), SWITCHOFFSET(MessageAt, '+09:00'), 120) AS MessageAt
                    FROM dbo.MeetingMessages
                    WHERE RoomCode = ? AND MessageID > ?
                    ORDER BY MessageID ASC
                    """,
                    room_code,
                    since_id,
                )
            rows = cursor.fetchall()

        messages = [
            {
                "message_id": row[0],
                "speaker_name": row[1],
                "language_code": row[2],
                "original_text": row[3],
                "translated_text": row[4],
                "message_at": row[5],
            }
            for row in rows
        ]
        return jsonify({"messages": messages})
    except Exception as exc:
        return jsonify({"error": f"failed to load messages: {exc}"}), 500


@app.post("/api/translate")
def translate_text():
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text", "")).strip()
    source_lang = str(payload.get("source", "")).strip().lower()
    target_lang = str(payload.get("target", "")).strip().lower()

    allowed = {"ko", "ja", "en"}
    if not text:
        return jsonify({"error": "text is required"}), 400
    if source_lang not in allowed or target_lang not in allowed:
        return jsonify({"error": "source/target must be ko, ja, or en"}), 400

    try:
        translator = GoogleTranslator(source=source_lang, target=target_lang)
        translated = translator.translate(text)
    except Exception as exc:
        return jsonify({"error": f"translation failed: {exc}"}), 500

    return jsonify({"translated_text": translated})


@app.get("/api/export-excel")
def export_excel():
    room_code = str(request.args.get("room_code", "")).strip()
    
    if not re.fullmatch(r"\d{6}", room_code):
        return jsonify({"error": "room_code must be 6 digits"}), 400
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            # 회의실 정보 조회
            cursor.execute(
                """
                SELECT TOP 1 RoomID, RoomName, RoomCode, 
                    CONVERT(VARCHAR(19), SWITCHOFFSET(CreatedAt, '+09:00'), 120) AS CreatedAt
                FROM dbo.MeetingRooms
                WHERE RoomCode = ? AND IsActive = 1
                ORDER BY RoomID DESC
                """,
                room_code,
            )
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "room not found"}), 404
            
            room_id = room_row[0]
            room_name = room_row[1] or "미정"
            created_at = room_row[3] or ""
            
            # 메시지 조회
            cursor.execute(
                """
                SELECT 
                    MessageID,
                    ISNULL(SpeakerName, '') AS SpeakerName,
                    LanguageCode,
                    OriginalText,
                    ISNULL(TranslatedText, '') AS TranslatedText,
                    CONVERT(VARCHAR(19), SWITCHOFFSET(MessageAt, '+09:00'), 120) AS MessageAt
                FROM dbo.MeetingMessages
                WHERE RoomCode = ?
                ORDER BY MessageID ASC
                """,
                room_code,
            )
            message_rows = cursor.fetchall()
        
        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "회의 기록"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 제목 영역
        ws['A1'] = f"회의실 #{room_code}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"회의실명: {room_name}"
        ws['A3'] = f"생성일: {created_at}"
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        
        # 컬럼 헤더
        headers = ["메시지ID", "발언자", "언어", "원문", "번역문", "시간"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        
        # 데이터 행
        center_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for row_num, msg_row in enumerate(message_rows, 6):
            msg_id, speaker_name, lang_code, original_text, translated_text, msg_at = msg_row
            
            ws.cell(row=row_num, column=1).value = msg_id
            ws.cell(row=row_num, column=2).value = speaker_name
            lang_name = "한국어" if lang_code == "ko" else ("English" if lang_code == "en" else "日本語")
            ws.cell(row=row_num, column=3).value = lang_name
            ws.cell(row=row_num, column=4).value = original_text
            ws.cell(row=row_num, column=5).value = translated_text
            ws.cell(row=row_num, column=6).value = msg_at
            
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = center_alignment
                cell.border = border
        
        # 엑셀 파일을 BytesIO에 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"회의_{room_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as exc:
        return jsonify({"error": f"failed to export excel: {exc}"}), 500


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5055))
    debug = os.getenv("DEBUG", "True").lower() in ("true", "1", "yes")
    app.run(host="0.0.0.0", port=port, debug=debug)
